"""
================================================================================
REFERENCIA: MOTOR PYTHON ORIGINAL (Google Colab) - A PORTAR A JAVA/SPRING BOOT
================================================================================
Este archivo NO se ejecuta. Es la especificacion funcional completa del
"motor" de Compras Directas que ya funciona en produccion (Google Colab +
Google Sheets) y que debe portarse a Java conservando EXACTAMENTE la misma
logica de negocio, cambiando solo la capa de datos (Google Sheets -> MySQL)
y la capa de E/S (Colab uploads -> formularios web Spring/Thymeleaf).

Cada funcion de aqui corresponde 1:1 a un metodo o servicio que debe existir
en el paquete com.granados.sistema.dafim.compras del proyecto Java.

ESQUEMA DE DATOS ACTUAL (Google Sheets -> se convierte en tablas MySQL):
# ---- ESQUEMA ACTUAL (Google Sheets) ----
HOJAS = {
    'contratos_029': ['nit','nombre','contrato','cargo','npg','anio'],
    'proveedores':   ['nit','nombre','renglon','descripcion'],
    'historial':     ['anio','mes','cheque','nit','nombre','renglon','monto','npg','modalidad','contrato','descripcion'],
}
"""

# ====================== MOTOR ======================

# ============================================================
#  SISTEMA DE COMPRAS DIRECTAS - MUNICIPALIDAD DE GRANADOS
#  DAFIM - Baja Verapaz | Art. 10 Num. 11 LAIP
#  Motor v2.0 - Match directo por numero de cheque
# ============================================================
import re, os
import pandas as pd
from collections import defaultdict, Counter

def norm(s):
    s = str(s).upper()
    for a,b in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U'),('Ü','U'),('Ñ','N')]:
        s = s.replace(a,b)
    return re.sub(r'[,\s]+', ' ', s).strip()

def limpiar_desc(d):
    d = str(d).strip()
    low = d.lower()
    for p in ['pago de ','por pago de ','pago  de ']:
        if low.startswith(p): d = d[len(p):]; break
    return d.strip()

def fmt_fecha(f):
    if hasattr(f,'strftime'): return f.strftime('%d/%m/%Y')
    s = str(f)
    if re.match(r'\d{4}-\d{2}-\d{2}', s): return s[8:10]+'/'+s[5:7]+'/'+s[:4]
    return s[:10]

def extraer_contrato(texto):
    """
    Extrae y NORMALIZA el numero de contrato de cualquier texto.
    Cubre: CONTRATO NO. 38-2026 | No: 038-2026 | numero 0038-2026 |
           NRO 38-2026 | N° 38 - 2026 | CONTRATO 38-2026 (sin nada)
    Normaliza ceros: 038-2026 / 0038-2026 -> 38-2026
    """
    t = str(texto).upper()
    m = re.search(
        r'CONTRATO\s*(?:N[UÚ]MERO|NRO|N[O°º]?)?\s*[.:;,]*\s*'
        r'0*(\d{1,3})\s*[\-–]\s*(\d{4})', t)
    if m:
        return f"{int(m.group(1))}-{m.group(2)}"
    return ''

def limpiar_nit(nit):
    return str(nit).strip().upper().replace('K','').replace(' ','').replace('.0','')

RENGLONES_CD = {'029','111','113','142','154','155','162','165','182','184','186',
                '187','188','191','196','199','211','223','232','241','262','263',
                '274','291','292','294','297','299','322','328','329','411'}
RENGLONES_EXCLUIDOS = {'011','015','022','035','051','055','062','063',
                       '331','332','413','422','435','769'}

KEYWORDS = {
    'BASURA':'142','BALASTO':'142','BALASTRO':'142','VIAJES DE MATERIAL':'142','FLETE':'142',
    'CAMIONADAS DE 10':'142','VIAJES DE ALUMNOS':'142','PIEDRIN':'142',
    'MAQUINARIA':'154','HORAS DE RENTA':'154','PATROL':'154','VIBROCOMPACTADOR':'154',
    'RETROEXCABADORA':'154','RETROEXCAVADORA':'154','TIPO PIPA':'154',
    'SACOS DE CEMENTO':'274','QUINTALES DE CEMENTO':'274','CEMENTO':'274',
    'CAMIONADAS DE ARENA':'223',
    'INSUMOS DE LIMPIEZA':'292','UTILES DE LIMPIEZA':'292',
    'INSUMOS DE LIBRERIA':'291','UTILES DE OFICINA':'291','LIBRERIA':'291',
    'PAPEL BOND':'291','HOJAS MEMBRETADAS':'291','TINTAS':'291',
    'VIVERES':'211','CANASTA BASICA':'211',
    'AUDIO Y SONIDO':'196','ACTIVIDADES DE FERIA':'196','FERIA':'196',
    'ALQUILER DE SILLAS':'196','SERVICIO DE SONIDO':'196','ALIMENTACION':'196',
    'PRESENTACION ARTISTICA':'187','MARIMBA':'187',
    'INTERNET':'186','FIBRA OPTICA':'186','ENERGIA ELECTRICA':'111',
    'CAJAS MORTUARIAS':'411','SERVICIOS FUNERARIOS':'411','FUNERARIO':'411',
    'ARRENDAMIENTO DE BIEN MUEBLE':'155',
    'RESTAURACION Y PINTURA':'165','MANTENIMIENTO Y REPARACION':'165',
    'BOMBILLAS':'297','FOTOCELDA':'297',
    'TOLDOS':'199','ESTUDIOS DE PROYECTOS':'199','CIRCUITO CERRADO':'199',
    'ALFOMBRA':'299','SEMILLAS':'299','MATERIALES VARIOS':'299',
    'MATERIALES UTILIZADOS':'299','CONOS':'299','IMPLEMENTOS':'299',
    'UNIFORME':'294','SERVICIOS PROFESIONALES':'188',
}
PALABRAS_COMUNES = {'GARCIA','GONZALEZ','MARTINEZ','REYES','LOPEZ','HERRERA',
                    'PEREZ','HERNANDEZ','RUIZ','MELGAR','ALVARADO','ROSALES',
                    'DE','LA','LOS','EL','Y','MAYEN','ORTIZ','GARCÍA'}
MESES_NOMBRE = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
                7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}

# ============ PARSERS ============
def parsear_cheques(path):
    """Reporte de cheques: col2=numero, col8=fecha, col14=estado, col15=nombre, col25=monto"""
    try:
        xls = pd.read_excel(path, sheet_name=0, engine='xlrd', header=None)
    except Exception:
        xls = pd.read_excel(path, sheet_name=0, header=None)
    cheques = []
    for _, row in xls.iterrows():
        num    = str(row[2]).strip().replace('.0','')  if pd.notna(row[2])  else ''
        estado = str(row[14]).strip() if len(row)>14 and pd.notna(row[14]) else ''
        if re.match(r'^\d{4,6}$', num) and estado == 'IMPRESO':
            cheques.append({'cheque':num,
                'nombre': str(row[15]).strip() if pd.notna(row[15]) else '',
                'monto':  round(float(row[25]),2) if pd.notna(row[25]) else 0.0,
                'fecha':  row[8]})
    if not cheques:
        raise ValueError("ERROR: no se encontraron cheques con estado IMPRESO. "
                         "Verifica que el archivo sea el reporte de cheques (.xls)")
    return cheques

MES_MAP = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
           'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}

def parsear_txt(path):
    """Bloques E56/E57/E58 de Guatecompras"""
    with open(path, encoding='utf-8', errors='replace') as f:
        content = f.read()
    blocks = [b.strip() for b in re.split(r'(?=E5[678]\d{7})', content.strip()) if b.strip()]
    regs, vistos = [], set()
    for b in blocks:
        lines = [l.strip() for l in b.split('\n') if l.strip()]
        if len(lines) < 8: continue
        npg = lines[0]
        if not re.match(r'^E5[678]\d{7}$', npg) or npg in vistos: continue
        vistos.add(npg)
        monto = 0.0
        for l in reversed(lines):
            t = l.replace(',','').strip()
            if re.match(r'^\d+\.?\d*$', t): monto = float(t); break
        linea6 = lines[6] if len(lines)>6 else ''
        modalidad = ('BAJA CUANTIA' if ('Art.43' in linea6 or 'Baja Cuant' in linea6)
                     else 'CASO DE EXCEPCION')
        nit, proveedor = '', ''
        for i,l in enumerate(lines[8:],8):
            if re.match(r'^\d+K?$', l, re.I):
                nit = limpiar_nit(l)
                proveedor = lines[i+1] if i+1<len(lines) else ''
                break
        m = re.match(r'(\d+)\.(\w+)\.(\d{4})', lines[1])
        fecha = (f"{m.group(1).zfill(2)}/{MES_MAP.get(m.group(2)[:3].lower(),'00')}/{m.group(3)}"
                 if m else lines[1])
        regs.append({'npg':npg,'fecha':fecha,'modalidad':modalidad,
                     'desc':lines[7] if len(lines)>7 else '',
                     'nit':nit,'proveedor':proveedor,'monto':round(monto,2)})
    return regs

def parsear_pdf(path):
    """
    Presupuesto SICOIN linea por linea. Cada registro incluye su NUMERO DE CHEQUE.
    Retorna: [{cheque, nit, nombre, renglon, monto, desc, contrato, status, fecha}]
    """
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        lineas = []
        for pg in pdf.pages:
            t = pg.extract_text() or ''
            lineas.extend(t.split('\n'))

    registros = []
    nit_actual, nombre_actual = '', ''
    reg_abierto = None

    RE_NIT  = re.compile(r'^(.{3,90}?)\s+NIT:\s*(E?\d+K?)\s*$')
    RE_REG  = re.compile(
        r'^(\d{2})\s(\d{2})\s(\d{3})\s(\d{3})\s(\d{3})\s(\d{3})\s+(\d+)\s*'
        r'(IMPRESO|ANULADO|CREADO)\s+.*?(\d{5})\s+(\d{2}/\d{2}/\d{4})\s+'
        r'(-?[\d,]+\.\d{2})\s*(.*)$')
    RE_FUENTE = re.compile(r'^\d{2}-\d{4}-\d{4}\s+\d{2}\s*$')
    RE_TOTAL  = re.compile(r'^Total por proveedor', re.I)
    RE_RUIDO  = re.compile(r'^(SIAF:|MUNICIPALIDAD|DEPARTAMENTO|CLASIFICACI|Usuario:|'
                           r'DETALLE DE|Ejercicio:|Nro Tipo|Estructura|Transac\.|'
                           r'P\xe1gina|Pagina|R008)', re.I)

    for ln in lineas:
        ln = ln.strip()
        if not ln: continue
        mnit = RE_NIT.match(ln)
        if mnit and 'NIT:' in ln:
            reg_abierto = None
            nombre_actual = mnit.group(1).strip()
            nit_raw = mnit.group(2)
            nit_actual = ('E'+nit_raw[1:] if nit_raw.startswith('E')
                          else nit_raw.upper().replace('K',''))
            continue
        mreg = RE_REG.match(ln)
        if mreg:
            renglon, status = mreg.group(6), mreg.group(8)
            cheque, fecha   = mreg.group(9), mreg.group(10)
            monto = float(mreg.group(11).replace(',',''))
            resto = mreg.group(12).strip()
            desc  = re.sub(r'^(Pago de|PAGO DE LA PLANILLA|PAGO DE)\s*','',resto).strip()
            reg = {'cheque':cheque,'nit':nit_actual,'nombre':nombre_actual,
                   'renglon':renglon,'monto':round(monto,2),'desc':desc,
                   'contrato':'N/A','status':status,'fecha':fecha}
            registros.append(reg)
            reg_abierto = reg
            continue
        if RE_TOTAL.match(ln): reg_abierto = None; continue
        if RE_FUENTE.match(ln) or RE_RUIDO.match(ln): continue
        if reg_abierto is not None:
            # continuar desc: texto normal, o linea que inicia con numero de contrato (49-2026)
            es_texto    = not re.match(r'^\d', ln)
            es_contrato = re.match(r'^\d{1,3}\-\d{4}\b', ln)
            if (es_texto or es_contrato) and len(reg_abierto['desc']) < 350:
                reg_abierto['desc'] += ' ' + ln

    for r in registros:
        r['desc'] = re.sub(r'\s+',' ', r['desc']).strip()
        c = extraer_contrato(r['desc'])
        if c: r['contrato'] = c
    return registros

# ============ PARSER: PDF DE CONFIRMACION NPG (Guatecompras) ============
# Se usa en el flujo "Cargar NPGs historicos". Distinto al PDF de SICOIN.
def parsear_pdf_npg(path):
    try:
        with pdfplumber.open(path) as pdf:
            texto = '\n'.join((pg.extract_text() or '') for pg in pdf.pages)
        m_npg  = re.search(r'Publicaci[oó]n\s*\(NPG\)[:\s]*([E]\d{8,10})', texto, re.I)
        m_nit  = re.search(r'Nit[:\s]+(\d+)\s*[-–]\s*(.+)', texto, re.I)
        m_desc = re.search(r'Descripci[oó]n[:\s]+(.+?)(?:Modalidad|$)', texto, re.I|re.DOTALL)
        desc   = re.sub(r'\s+',' ',(m_desc.group(1) if m_desc else '')).strip()
        return {'npg': m_npg.group(1).strip() if m_npg else '',
                'nit': m_nit.group(1).strip() if m_nit else '',
                'nombre': m_nit.group(2).strip() if m_nit else '',
                'contrato': _con_txt(desc) or _con_doc(texto), 'desc': desc[:200]}
    except Exception as e:
        return {'npg':'','nit':'','nombre':'','contrato':'','desc':'','error':str(e)}


def parsear_remuneraciones(path):
    df = pd.read_excel(path, header=None)
    personas = []
    for _, row in df.iterrows():
        try:
            if int(float(str(row[1]))) == 29:
                nombre = str(row[2]).strip() if pd.notna(row[2]) else ''
                monto  = float(row[6]) if pd.notna(row[6]) else 0
                cargo  = str(row[3]).strip() if pd.notna(row[3]) else ''
                if nombre and nombre != 'nan' and monto > 0:
                    personas.append({'nombre':nombre,'monto':round(monto,2),'cargo':cargo})
        except Exception: pass
    return personas


# ============ PARSER: MACHOTE (Excel de compras directas ya hecho) ============
def parsear_machote(path):
    """
    Lee un Excel de compras directas de cualquier mes (formato 15 columnas).
    Infiere renglon por keywords si la columna esta vacia.
    """
    df = pd.read_excel(path, header=None)
    inicio = None
    for i in range(min(40, df.shape[0])):
        v = str(df.iloc[i,0]).strip() if pd.notna(df.iloc[i,0]) else ''
        if v in ('1','1.0'): inicio = i; break
    if inicio is None:
        raise ValueError("No se encontro la primera fila de datos (columna A = 1). "
                         "Verifica que sea un Excel de compras directas.")
    def reng_kw(texto):
        t = str(texto).upper()
        for kw in sorted(KEYWORDS, key=len, reverse=True):
            if kw in t: return KEYWORDS[kw]
        return ''
    filas = []
    for i in range(inicio, df.shape[0]):
        row = [str(v).strip() if pd.notna(v) else '' for v in df.iloc[i]]
        if not row[0] or not re.match(r'^\d+(\.0)?$', row[0]): break
        while len(row) < 15: row.append('')
        renglon = row[6].replace('.0','').strip()
        if renglon and renglon.isdigit(): renglon = renglon.zfill(3)
        if not renglon:
            # inferir: si la desc menciona contrato de servicios -> 029, sino keywords
            if extraer_contrato(row[2]) and 'SERVICIOS' in row[2].upper():
                renglon = '029'
            else:
                renglon = reng_kw(row[2]) or '299'
        try: monto = float(str(df.iloc[i,4]).replace(',','')) if pd.notna(df.iloc[i,4]) else 0.0
        except Exception: monto = 0.0
        contrato = row[13].replace('.0','').strip()
        if contrato in ('','0','nan','N/A','NA'):
            contrato = extraer_contrato(row[2]) or 'N/A'
        else:
            # normalizar el de la columna: 038-2026 -> 38-2026
            mn = re.match(r'^0*(\d{1,3})\s*[\-–]\s*(\d{4})$', contrato)
            if mn: contrato = f"{int(mn.group(1))}-{mn.group(2)}"
        npg = row[9].strip().upper()
        if npg and not re.match(r'^E\d{8,10}$', npg): npg = ''
        filas.append({'modalidad':row[1] or 'CASO DE EXCEPCION','desc':row[2],
                      'precio':round(monto,2),'renglon':renglon,'proveedor':row[7],
                      'nit':limpiar_nit(row[8]),'npg':npg,'fecha_pub':row[10],
                      'fecha_adj':row[11],'contrato':contrato,
                      'fecha_cont':row[14] or 'N/A','cheque':''})
    return filas

def extraer_para_bd(filas):
    """De una lista de filas (machote o procesadas) extrae datos para la BD."""
    nuevos_029, nuevos_prov = {}, {}
    for f in filas:
        nit = f.get('nit','')
        if not nit: continue
        if f['renglon'] == '029':
            cm = re.search(r'COMO[:\s]+(.{4,60}?)[,.]', str(f['desc']).upper())
            d = nuevos_029.setdefault(nit, {'nombre':f['proveedor'],'contrato':'N/A',
                                            'cargo':'','npg':''})
            if f['contrato'] != 'N/A': d['contrato'] = f['contrato']
            if cm and not d['cargo']: d['cargo'] = cm.group(1).strip()
            if f.get('npg') and not d['npg']: d['npg'] = f['npg']
        elif f['renglon'] not in ('184',''):
            if nit not in nuevos_prov:
                nuevos_prov[nit] = {'nombre':f['proveedor'],'renglon':f['renglon'],
                                    'desc':limpiar_desc(str(f['desc']))[:120]}
    return nuevos_029, nuevos_prov

# ============ MOTOR PRINCIPAL ============
# ============================================================================
# NOTA PORTEO (CRITICO): Esta es LA funcion mas importante del sistema.
# El orden de prioridad para resolver cada cheque es:
#   1) Match directo por numero de cheque contra el PDF SICOIN (fuente de
#      verdad: ahi viene nit, renglon, descripcion y contrato).
#   2) Si el PDF no tiene ese cheque -> buscar por nombre difuso en
#      BD_CONTRATOS (personal 029): interseccion de palabras >= 3 Y al
#      menos 1 palabra "significativa" (que no sea apellido comun/articulo).
#   3) Si tampoco -> buscar por nombre difuso en BD_PROVEEDORES (>= 3
#      palabras en comun).
#   4) Si tampoco -> inferir renglon por KEYWORDS en la descripcion, o
#      '299' por defecto, y generar una ALERTA "REVISAR".
# El NPG y la modalidad se resuelven distinto si renglon == '029'
# (personal, modalidad siempre CASO DE EXCEPCION, NPG viene de la BD
# historica) vs otros renglones (NPG/modalidad del TXT de Guatecompras del
# mes, con anti-duplicado global por NPG).
# Al final, las filas NO-029 van primero y las 029 al final del Excel.
# ============================================================================
def construir_filas(cheques, txt_regs, pdf_regs, bd_proveedores=None, bd_contratos=None, bd_npgs_029=None):
    """
    Match DIRECTO por numero de cheque (PDF tiene el cheque en cada registro).
    bd_proveedores: {NIT: {nombre, renglon, desc}}      - opcional
    bd_contratos:   {NIT: {nombre, contrato, cargo}}    - opcional
    bd_npgs_029:    {NIT: npg}                           - NPGs historicos R029
    """
    bd_proveedores = bd_proveedores or {}
    bd_contratos   = bd_contratos or {}
    bd_npgs_029    = bd_npgs_029 or {}

    # Indice PDF por cheque (solo IMPRESO y renglones CD)
    pdf_por_cheque = defaultdict(list)
    for r in pdf_regs:
        if r['status']=='IMPRESO' and r['renglon'] in RENGLONES_CD:
            pdf_por_cheque[r['cheque']].append(r)

    # Pool TXT por (nit, monto) y por nit
    txt_pool = defaultdict(list)
    nit_npgs_txt = defaultdict(list)
    for r in txt_regs:
        if r['nit'] and r['monto']>0:
            txt_pool[(r['nit'], r['monto'])].append(dict(r))
        if r['nit'] and r['npg']:
            nit_npgs_txt[r['nit']].append(r)

    def reng_kw(texto):
        t = texto.upper()
        for kw in sorted(KEYWORDS, key=len, reverse=True):
            if kw in t: return KEYWORDS[kw]
        return ''

    filas, alertas = [], []
    npg_usados = set()
    nuevos_029, nuevos_prov = {}, {}

    for c in cheques:
        nombre_c, monto_c = c['nombre'], c['monto']
        fecha_c, num_cheq = fmt_fecha(c['fecha']), c['cheque']
        renglon=''; nit=''; npg=''; contrato='N/A'
        desc=''; proveedor=nombre_c; modalidad=''

        # === MATCH DIRECTO POR CHEQUE (la fuente de verdad) ===
        regs_pdf = pdf_por_cheque.get(num_cheq, [])
        if regs_pdf:
            # Un cheque puede tener varias lineas presupuestarias; tomar la de mayor monto
            r = max(regs_pdf, key=lambda x: x['monto'])
            nit, renglon = r['nit'], r['renglon']
            desc = r['desc']
            if r['contrato'] != 'N/A': contrato = r['contrato']

        # === Si el PDF no lo tiene (cheque sin registro CD) ===
        if not renglon:
            # BD de contratos (R029 conocidos)
            palabras = set(norm(nombre_c).split())
            sig = palabras - PALABRAS_COMUNES
            for bnit, fila in bd_contratos.items():
                pn = set(norm(fila['nombre']).split())
                if len(palabras & pn) >= 3 and len(sig & (pn-PALABRAS_COMUNES)) >= 1:
                    nit, renglon = bnit, '029'
                    contrato = fila.get('contrato','N/A')
                    cargo = fila.get('cargo','')
                    if cargo:
                        desc = (f"POR PAGO DE SERVICIOS PRESTADOS A LA MUNICIPALIDAD DE "
                                f"GRANADOS COMO: {cargo}, SEGUN CONTRATO NO. {contrato}")
                    break
        if not renglon:
            palabras = set(norm(nombre_c).split())
            for bnit, fila in bd_proveedores.items():
                pn = set(norm(fila['nombre']).split())
                if len(palabras & pn) >= 3:
                    nit, renglon = bnit, fila['renglon']
                    if fila.get('desc'): desc = fila['desc']
                    break
        if not renglon:
            renglon = reng_kw(nombre_c) or '299'
            alertas.append(f"REVISAR cheque {num_cheq}: {nombre_c} | Q{monto_c:,.2f} "
                           f"-> R{renglon} sin registro en PDF")

        # === MODALIDAD Y NPG desde el TXT ===
        if renglon == '029':
            modalidad = 'CASO DE EXCEPCION'
            # NPG R029: reutilizar de BD historica, o del TXT si publico este mes
            if nit in bd_npgs_029:
                npg = bd_npgs_029[nit]
            else:
                kt = (nit, monto_c)
                if txt_pool[kt]:
                    t = txt_pool[kt].pop(0); npg = t['npg']
        else:
            # NPG no-029: SIEMPRE del TXT del mes (publicacion nueva)
            kt = (nit, monto_c)
            if txt_pool[kt]:
                t = txt_pool[kt].pop(0)
                npg = t['npg']; modalidad = t['modalidad']
                if t['desc'] and len(t['desc']) > len(desc): pass  # PDF desc es mejor
            else:
                # respaldo: cualquier NPG del NIT no usado
                for t in nit_npgs_txt.get(nit, []):
                    if t['npg'] not in npg_usados:
                        npg = t['npg']; modalidad = t['modalidad']; break
            if not modalidad:
                modalidad = 'BAJA CUANTIA' if monto_c <= 25000 else 'CASO DE EXCEPCION'

        # Anti-duplicado NPG
        if npg:
            if npg in npg_usados: npg = ''
            else: npg_usados.add(npg)

        if renglon=='029' and contrato=='N/A':
            c2 = extraer_contrato(desc)
            if c2: contrato = c2
            else: alertas.append(f"R029 SIN CONTRATO cheque {num_cheq}: {nombre_c}")
        if renglon=='029' and nit and nit not in bd_contratos:
            alertas.append(f"🆕 R029 NUEVO (no esta en BD): {nombre_c} | NIT:{nit} "
                           f"| contrato:{contrato} -> se agregara al guardar")
        if not nit:
            alertas.append(f"SIN NIT cheque {num_cheq}: {nombre_c} | Q{monto_c:,.2f}")

        if not desc: desc = nombre_c
        filas.append({'cheque':num_cheq,'modalidad':modalidad,'desc':limpiar_desc(desc),
                      'precio':monto_c,'renglon':renglon,'proveedor':proveedor,
                      'nit':nit,'npg':npg,'fecha_pub':fecha_c,'fecha_adj':fecha_c,
                      'contrato':contrato,'fecha_cont':'N/A'})

        # Acumular para la BD
        if renglon=='029' and nit:
            cm = re.search(r'COMO[:\s]+(.{4,60}?)[,.]', desc.upper())
            nuevos_029[nit] = {'nombre':proveedor,'contrato':contrato,
                               'cargo': cm.group(1).strip() if cm else '',
                               'npg': npg or bd_npgs_029.get(nit,'')}
        elif nit and renglon not in ('029','184'):
            nuevos_prov[nit] = {'nombre':proveedor,'renglon':renglon,
                                'desc':limpiar_desc(desc)[:120]}

    filas = [f for f in filas if f['renglon']!='029'] + [f for f in filas if f['renglon']=='029']
    return filas, alertas, nuevos_029, nuevos_prov


# ============ VALIDACION CRUZADA CON REMUNERACIONES ============
def validar_remuneraciones(filas, personas_rem):
    """
    Compara los R029 del Excel contra el archivo de remuneraciones del mes.
    Retorna lista de observaciones (no errores fatales).
    """
    obs = []
    if not personas_rem:
        return ["(sin archivo de remuneraciones - validacion omitida)"]
    filas_029 = [f for f in filas if f['renglon'] == '029']
    monto_filas = round(sum(f['precio'] for f in filas_029), 2)
    monto_rem   = round(sum(p['monto'] for p in personas_rem), 2)
    obs.append(f"R029 en Excel: {len(filas_029)} pagos | Q{monto_filas:,.2f}")
    obs.append(f"R029 en remuneraciones: {len(personas_rem)} personas | Q{monto_rem:,.2f}")
    if abs(monto_filas - monto_rem) < 0.01:
        obs.append("[OK] Montos R029 cuadran con remuneraciones")
    else:
        obs.append(f"[REV] Diferencia de Q{abs(monto_filas-monto_rem):,.2f} "
                   "(puede ser pago de mes atrasado o persona en otra planilla)")
    # personas en remuneraciones que no aparecen en el Excel
    nombres_excel = {norm(f['proveedor']) for f in filas_029}
    faltantes = []
    for p in personas_rem:
        pn = set(norm(p['nombre']).split())
        if not any(len(pn & set(ne.split())) >= 2 for ne in nombres_excel):
            faltantes.append(p['nombre'])
    if faltantes:
        obs.append(f"[REV] En remuneraciones pero SIN cheque este mes ({len(faltantes)}):")
        for n in faltantes[:10]: obs.append(f"      - {n}")
    else:
        obs.append("[OK] Todas las personas de remuneraciones tienen su cheque")
    return obs

# ============ VALIDADOR ============
def validar(filas, cheques):
    rep, ok = [], True
    n_ch, n_f = len(cheques), len(filas)
    m_ch = round(sum(c['monto'] for c in cheques),2)
    m_f  = round(sum(f['precio'] for f in filas),2)
    def chk(cond, msg_ok, msg_err):
        nonlocal ok
        rep.append(("[OK]  " if cond else "[ERR] ") + (msg_ok if cond else msg_err))
        if not cond: ok = False
    chk(n_f==n_ch, f"Filas {n_f} = cheques IMPRESO {n_ch}",
        f"Filas {n_f} != cheques {n_ch}")
    chk(abs(m_f-m_ch)<0.01, f"Monto cuadra al centavo: Q{m_f:,.2f}",
        f"Monto NO cuadra: filas Q{m_f:,.2f} vs cheques Q{m_ch:,.2f}")
    dups=[n for n,c in Counter(f['npg'] for f in filas if f['npg']).items() if c>1]
    chk(not dups, "Sin NPGs duplicados", f"NPGs duplicados: {dups[:5]}")
    sin_reng=[f for f in filas if not f['renglon']]
    chk(not sin_reng, "Todas las filas tienen renglon", f"{len(sin_reng)} sin renglon")
    malos=[f for f in filas if f['renglon'] in RENGLONES_EXCLUIDOS]
    chk(not malos, "Sin renglones excluidos", f"{len(malos)} con renglon excluido")
    sin_nit=[f for f in filas if not f['nit']]
    rep.append(f"[{'OK ' if not sin_nit else 'REV'}]  Filas sin NIT: {len(sin_nit)}")
    r029=[f for f in filas if f['renglon']=='029']
    sc=[f for f in r029 if f['contrato'] in ('N/A','')]
    rep.append(f"[{'OK ' if not sc else 'REV'}]  R029: {len(r029)} | sin contrato: {len(sc)}")
    cnt=Counter(f['renglon'] for f in filas)
    rep.append("       Distribucion: " + ', '.join(f"R{k}:{v}" for k,v in sorted(cnt.items())))
    rep.append(f"       BAJA CUANTIA: {sum(1 for f in filas if f['modalidad']=='BAJA CUANTIA')} "
               f"| CASO EXCEPCION: {sum(1 for f in filas if f['modalidad']!='BAJA CUANTIA')}")
    return ok, rep


# ============ REPORTE DE DIFERENCIAS VS MES ANTERIOR ============
def comparar_con_mes_anterior(filas, historial, anio, mes):
    """
    Compara el mes procesado contra el mes anterior guardado en el historial.
    historial: lista de dicts con keys anio, mes, nit, nombre, monto, renglon
    """
    obs = []
    mes_ant, anio_ant = (mes-1, anio) if mes > 1 else (12, anio-1)
    prev = [h for h in historial
            if str(h.get('anio'))==str(anio_ant) and str(h.get('mes'))==str(mes_ant)]
    if not prev:
        return [f"(no hay datos de {mes_ant}/{anio_ant} en la BD para comparar)"]
    obs.append(f"Comparando contra {mes_ant:02d}/{anio_ant} ({len(prev)} registros):")
    nits_prev = {str(h.get('nit','')) for h in prev if h.get('nit')}
    nits_hoy  = {f['nit'] for f in filas if f['nit']}
    nuevos    = nits_hoy - nits_prev
    se_fueron = nits_prev - nits_hoy
    if nuevos:
        obs.append(f"  🆕 Proveedores nuevos este mes: {len(nuevos)}")
        nombres = {f['nit']: f['proveedor'] for f in filas}
        for n in list(nuevos)[:8]: obs.append(f"     + {nombres.get(n,'')[:45]} ({n})")
        if len(nuevos) > 8: obs.append(f"     ... y {len(nuevos)-8} mas")
    if se_fueron:
        obs.append(f"  👋 Cobraron el mes pasado pero NO este mes: {len(se_fueron)}")
        nombres_prev = {str(h.get('nit','')): h.get('nombre','') for h in prev}
        for n in list(se_fueron)[:8]: obs.append(f"     - {nombres_prev.get(n,'')[:45]} ({n})")
        if len(se_fueron) > 8: obs.append(f"     ... y {len(se_fueron)-8} mas")
    # cambios de monto en R029
    monto_prev_029 = {}
    for h in prev:
        if str(h.get('renglon','')).zfill(3) == '029':
            try: monto_prev_029[str(h['nit'])] = float(h.get('monto',0))
            except Exception: pass
    cambios = []
    for f in filas:
        if f['renglon']=='029' and f['nit'] in monto_prev_029:
            if abs(f['precio'] - monto_prev_029[f['nit']]) > 0.01:
                cambios.append((f['proveedor'], monto_prev_029[f['nit']], f['precio']))
    if cambios:
        obs.append(f"  💱 R029 con cambio de monto: {len(cambios)}")
        for nom, antes, ahora in cambios[:8]:
            obs.append(f"     {nom[:35]}: Q{antes:,.2f} -> Q{ahora:,.2f}")
    if not nuevos and not se_fueron and not cambios:
        obs.append("  [OK] Sin diferencias relevantes")
    return obs

# ============ EXCEL ============
# ============================================================================
# NOTA PORTEO: El formato de este Excel es un ENTREGABLE LEGAL (informa al
# amparo del Art. 10 Num. 11 LAIP). Replicar EXACTO con Apache POI:
#  - 6 filas de titulo combinadas (merge A:O), negrita, centrado.
#  - Encabezados en fila 9 (15 columnas), con bordes.
#  - Columnas de texto forzado (number_format '@'): 2,3,7,8,9,10,11,12,13,14,15
#    (para que NIT/NPG/contrato/renglon no se conviertan en numero/fecha).
#  - Columna F (MONTO TOTAL) = formula "=D*E" por fila.
#  - Fila final: "TOTAL" en col 5, "=SUM(F.. :F..)" en col 6.
#  - Anchos de columna y alineaciones tal como cw/al en el codigo.
# ============================================================================
def generar_excel(filas, mes, anio, path_salida):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='COMPRAS DIRECTAS'
    thin=Side(style='thin'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    fd=Font(name='Calibri',size=9); fb=Font(name='Calibri',size=9,bold=True)
    ft=Font(name='Calibri',size=11,bold=True)
    ac=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ai=Alignment(horizontal='left',vertical='center',wrap_text=True)
    ad=Alignment(horizontal='right',vertical='center')
    titulos=['MUNICIPALIDAD DE GRANADOS, BAJA VERAPAZ',
             'DIRECCION DE ADMINISTRACION FINANCIERA INTEGRADA MUNICIPAL -DAFIM-',
             'CLASIFICACION INSTITUCIONAL SIAF: 12101505',
             'INFORMACION DE OFICIO, ARTICULO 10, NUMERAL 11, LEY DE ACCESO A LA INFORMACION PUBLICA',
             'COMPRAS DIRECTAS / CONTRATACIONES DE BIENES Y SERVICIOS',
             f'CORRESPONDIENTE AL MES DE {MESES_NOMBRE[mes]} {anio}']
    for i,t in enumerate(titulos,1):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=15)
        c=ws.cell(row=i,column=1,value=t); c.font=ft; c.alignment=ac
    hdr=['No.','MODALIDAD DE COMPRA','DESCRIPCION DE LA COMPRA O CONTRATACION','CANTIDAD',
         'PRECIO UNITARIO','MONTO TOTAL DE LA COMPRA','RENGLON','NOMBRE DEL PROVEEDOR',
         'NIT','NPG','FECHA PUBLICACION','FECHA ADJUDICACION','ACTIVA',
         'NUMERO DE CONTRATO','FECHA DE CONTRATO']
    hr=9
    for ci,h in enumerate(hdr,1):
        c=ws.cell(row=hr,column=ci,value=h); c.font=fb; c.alignment=ac; c.border=border
    cw={1:5,2:18,3:55,4:6,5:13,6:13,7:8,8:30,9:13,10:14,11:13,12:13,13:10,14:15,15:13}
    al=[ac,ac,ai,ac,ad,ad,ac,ai,ac,ac,ac,ac,ac,ac,ac]
    for col,w in cw.items(): ws.column_dimensions[get_column_letter(col)].width=w
    ds=hr+1
    for idx,f in enumerate(filas,1):
        rn=ds+idx-1; ws.row_dimensions[rn].height=30
        vals=[idx,f['modalidad'],f['desc'],1,float(f['precio']),f'=+D{rn}*E{rn}',
              f['renglon'],f['proveedor'],f['nit'],f['npg'],f['fecha_pub'],
              f['fecha_adj'],'ACTIVA',f['contrato'],f['fecha_cont']]
        # Columnas que DEBEN ser texto para que Excel no las convierta
        # (contrato 12-2026 -> diciembre, renglon 029 -> 29, NIT/NPG largos)
        COLS_TEXTO = {2,3,7,8,9,10,11,12,13,14,15}
        for ci,(v,a) in enumerate(zip(vals,al),1):
            c=ws.cell(row=rn,column=ci,value=str(v) if ci in COLS_TEXTO else v)
            c.font=fd; c.alignment=a; c.border=border
            if ci in (5,6): c.number_format='#,##0.00'
            elif ci in COLS_TEXTO: c.number_format='@'
    tr=ds+len(filas)
    for ci in range(1,16): ws.cell(row=tr,column=ci).border=border
    l=ws.cell(row=tr,column=5,value='TOTAL'); l.font=fb; l.alignment=ad
    t=ws.cell(row=tr,column=6,value=f'=SUM(F{ds}:F{tr-1})')
    t.font=fb; t.number_format='#,##0.00'; t.alignment=ad
    wb.save(path_salida)
    return path_salida



# ============================================================================
# OPERACIONES DE BASE DE DATOS (Google Sheets) - REFERENCIA DE COMPORTAMIENTO
# ============================================================================
# Estas funciones NO se portan tal cual (Google Sheets -> MySQL/JPA), pero
# definen el COMPORTAMIENTO esperado de la capa de persistencia. Cada una
# debe convertirse en metodo(s) de Repository/Service usando JPA:
#
#   cargar_bd()        -> ContratoNit029Repository.findAll() +
#                          ProveedorRepository.findAll() (mapear a Map<nit,...>)
#   cargar_historial() -> HistorialCompraRepository.findAll()
#   guardar_bd_batch() -> guardar/actualizar (upsert) en contratos_029 y
#                          proveedores. OJO: en Sheets se hace ws.clear()+
#                          reescritura total; en MySQL debe ser UPSERT real
#                          por NIT (no borrar toda la tabla).
#   guardar_historial()-> primero DELETE de los registros existentes con
#                          ese (anio, mes), luego INSERT de las filas nuevas
#                          (es lo que permite re-procesar un mes sin duplicar).
#   fusionar_bd()       -> logica de merge in-memory entre lo que ya esta en
#                          BD y lo nuevo calculado por construir_filas(). Esta
#                          logica de merge SI debe portarse igual, pero
#                          operando sobre entidades JPA en vez de dicts.
# ============================================================================

def cargar_bd():
    bd_c, bd_n, bd_p = {}, {}, {}
    for rec in sh.worksheet('contratos_029').get_all_records():
        nit = str(rec.get('nit','')).strip()
        if nit:
            bd_c[nit] = {'nombre':str(rec.get('nombre','')),
                         'contrato':str(rec.get('contrato','N/A')) or 'N/A',
                         'cargo':str(rec.get('cargo','')),
                         'npg':str(rec.get('npg','')).strip(),
                         'anio':rec.get('anio','')}
            if bd_c[nit]['npg']: bd_n[nit] = bd_c[nit]['npg']
    for rec in sh.worksheet('proveedores').get_all_records():
        nit = str(rec.get('nit','')).strip()
        if nit:
            bd_p[nit] = {'nombre':str(rec.get('nombre','')),
                         'renglon':str(rec.get('renglon','')).zfill(3),
                         'desc':str(rec.get('descripcion',''))}
    return bd_c, bd_n, bd_p

def cargar_historial():
    return sh.worksheet('historial').get_all_records()

def guardar_bd_batch(bd_c, bd_p):
    ws = sh.worksheet('contratos_029')
    filas = [['nit','nombre','contrato','cargo','npg','anio']]
    filas += [[n, d['nombre'], d['contrato'], d['cargo'], d['npg'], d.get('anio','')]
              for n, d in sorted(bd_c.items())]
    ws.clear(); ws.update('A1', filas)
    ws2 = sh.worksheet('proveedores')
    filas2 = [['nit','nombre','renglon','descripcion']]
    filas2 += [[n, d['nombre'], d['renglon'], d['desc']] for n, d in sorted(bd_p.items())]
    ws2.clear(); ws2.update('A1', filas2)

def guardar_historial(filas_mes, anio, mes):
    ws = sh.worksheet('historial')
    todos = ws.get_all_values()
    encabezado = todos[0] if todos else HOJAS['historial']
    conservar = [r for r in todos[1:]
                 if not (len(r)>1 and str(r[0])==str(anio) and str(r[1])==str(mes))]
    reemplazadas = len(todos) - 1 - len(conservar)
    nuevas = [[anio, mes, f.get('cheque',''), f['nit'], f['proveedor'], f['renglon'],
               f['precio'], f['npg'], f['modalidad'], f['contrato'],
               str(f['desc'])[:200]] for f in filas_mes]
    ws.clear()
    datos = [encabezado] + conservar + nuevas
    for i in range(0, len(datos), 500):
        if i == 0: ws.update('A1', datos[:500])
        else: ws.append_rows(datos[i:i+500]); time.sleep(1.0)
    return reemplazadas

def fusionar_bd(bd_c, bd_p, n029, nprov, anio):
    for nit, d in n029.items():
        if nit in bd_c:
            ex = bd_c[nit]
            if d['contrato'] != 'N/A': ex['contrato'] = d['contrato']
            if d['cargo'] and not ex['cargo']: ex['cargo'] = d['cargo']
            if d.get('npg'): ex['npg'] = d['npg']
            ex['anio'] = anio
        else:
            bd_c[nit] = {'nombre':d['nombre'],'contrato':d['contrato'],
                         'cargo':d['cargo'],'npg':d.get('npg',''),'anio':anio}
    for nit, d in nprov.items():
        if nit not in bd_p: bd_p[nit] = dict(d)
